from openpyxl import Workbook
from collections import defaultdict

def export_check(text):
    wb = Workbook()
    checks = text.strip().split('---')
    
    for check_index, check in enumerate(checks):
        ws = wb.create_sheet(title=f'Чек {check_index + 1}')
        products = defaultdict(lambda: {'price': 0, 'quantity': 0})
        lines = check.strip().split('\n')
        for line in lines:
            parts = line.split('\t')
            if len(parts) != 3:
                continue
            
            item_name = parts[0]
            item_price = float(parts[1])
            item_quantity = int(parts[2])
            
            key = (item_name, item_price)
            products[key]['quantity'] += item_quantity
        
        sorted_products = sorted(products.items(), key=lambda x: x[0][0])
        
        ws.append(["Товар", "Цена за единицу", "Количество", "Общая стоимость"])
        
        for (item_name, item_price), details in sorted_products:
            item_quantity = details['quantity']
            total_price = item_price * item_quantity
            
            ws.append([item_name, item_price, item_quantity, total_price])
        
        total_row_index = len(sorted_products) + 2  
        ws.append(["Итого", "", "", f"=SUM(D2:D{total_row_index - 1})"]) 
    
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    wb.save('res3.xlsx')
    print("Файл 'res.xlsx' успешно создан!")


text = """Яблоки\t50.5\t10
Бананы\t25.0\t5
Груши\t35.75\t8
Апельсины\t40.0\t2
---
Яблоки\t50.5\t3
Апельсины\t40.0\t1
Бананы\t25.0\t7"""

export_check(text)
