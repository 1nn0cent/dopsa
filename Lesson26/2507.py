from openpyxl import Workbook

def export_check(text):
    wb = Workbook()
    ws = wb.active
    lines = text.strip().split('\n')
    
    for i, line in enumerate(lines):
        parts = line.split('\t')
        if len(parts) != 3:
            continue
        
        item_name = parts[0]
        item_price = float(parts[1])
        item_quantity = int(parts[2])
 
        ws.cell(row=i+1, column=1, value=item_name)  
        ws.cell(row=i+1, column=2, value=item_price)  
        ws.cell(row=i+1, column=3, value=item_quantity)  
        
        ws.cell(row=i+1, column=4, value=f"=B{i+1}*C{i+1}")  
    
    total_row = len(lines) + 1
    ws.cell(row=total_row, column=1, value="Итого")
    ws.cell(row=total_row, column=4, value=f"=SUM(D1:D{total_row-1})")  
    wb.save('res2.xlsx')

text = """
Яблоки\t50.5\t10
Бананы\t25.0\t5
Груши\t35.75\t8
Апельсины\t40.0\t2
"""

export_check(text)

